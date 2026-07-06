import os
import shutil
from datetime import date, datetime, timezone, timedelta
from decimal import Decimal
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.user import User
from app.models.financial import MJOPUpload, MJOPItem, Quote, ReserveFondsEntry, ContributionPlan
from app.core.dependencies import get_current_user, get_current_beheerder, require_vve_access
from app.core.config import settings
from app.schemas.financial import (
    MJOPUploadOut, MJOPItemOut, MJOPItemCreate, MJOPItemUpdate,
    QuoteCreate, QuoteApprove, QuoteOut,
    ReserveFondsEntryCreate, ReserveFondsEntryOut,
    ContributionPlanCreate, ContributionPlanOut,
)
from app.services.mjop_parser import parse_mjop_file
from app.services.balanssheet_parser import parse_balanssheet
from app.services.scenario_calculator import (
    FinancialInput, MJOPItemInput, calculate_all_scenarios, smart_planning
)

router = APIRouter(prefix="/api/vves/{vve_id}/financial", tags=["financial"])


def _bereken_auto_bijdragen(
    last_entry_date: date,
    today: date,
    amount_per_period: Decimal,
    frequency: str,
) -> Decimal:
    """Tel bijdragen mee die automatisch geboekt worden (20ste van elke periode)."""
    if frequency == "monthly":
        count = 0
        year, month = last_entry_date.year, last_entry_date.month
        while True:
            booking = date(year, month, 20)
            if last_entry_date < booking <= today:
                count += 1
            if month == 12:
                year, month = year + 1, 1
            else:
                month += 1
            if date(year, month, 1) > today:
                break
        return amount_per_period * count
    elif frequency == "quarterly":
        count = 0
        for year in range(last_entry_date.year, today.year + 1):
            for month in (1, 4, 7, 10):
                booking = date(year, month, 20)
                if last_entry_date < booking <= today:
                    count += 1
        return amount_per_period * count
    return Decimal(0)


def _check_vve_access(vve_id: int, user: User, db: Session):
    require_vve_access(vve_id, user, db)


# --- MJOP Upload ---

@router.post("/mjop/upload", response_model=MJOPUploadOut, status_code=status.HTTP_201_CREATED)
async def upload_mjop(
    vve_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_beheerder),
    db: Session = Depends(get_db),
):
    _check_vve_access(vve_id, current_user, db)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".xlsx", ".xls", ".xlsm", ".pdf"):
        raise HTTPException(status_code=400, detail="Alleen Excel (.xlsx, .xls) en PDF bestanden zijn toegestaan")

    upload_dir = os.path.join(settings.UPLOAD_DIR, "mjop", str(vve_id))
    os.makedirs(upload_dir, exist_ok=True)
    stored_name = f"{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    file_path = os.path.join(upload_dir, stored_name)

    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    upload = MJOPUpload(
        vve_id=vve_id,
        original_filename=file.filename,
        stored_filename=stored_name,
        status="processing",
        uploaded_by_id=current_user.id,
    )
    db.add(upload)
    db.commit()
    db.refresh(upload)

    # Synchroon parsen zodat fouten direct als HTTP-fout terugkomen
    try:
        items = parse_mjop_file(file_path)
        if not items:
            upload.status = "failed"
            db.commit()
            raise HTTPException(status_code=422, detail="Geen posten gevonden in het bestand. Controleer of het bestand een MJOP bevat met jaar, omschrijving en bedrag.")
        for item_data in items:
            db.add(MJOPItem(vve_id=vve_id, mjop_upload_id=upload.id, **item_data))
        upload.status = "active"
        db.query(MJOPUpload).filter(
            MJOPUpload.vve_id == vve_id,
            MJOPUpload.id != upload.id,
            MJOPUpload.status == "active",
        ).update({"status": "archived"})
        db.commit()
        db.refresh(upload)
    except HTTPException:
        raise
    except Exception as exc:
        upload.status = "failed"
        db.commit()
        raise HTTPException(status_code=500, detail=f"Fout bij verwerken: {exc}")

    return upload


@router.get("/mjop/uploads", response_model=list[MJOPUploadOut])
def list_mjop_uploads(vve_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _check_vve_access(vve_id, current_user, db)
    return db.query(MJOPUpload).filter(MJOPUpload.vve_id == vve_id).order_by(MJOPUpload.uploaded_at.desc()).all()


# --- MJOP Items ---

@router.get("/mjop/items", response_model=list[MJOPItemOut])
def list_mjop_items(vve_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _check_vve_access(vve_id, current_user, db)
    # Actieve upload items ophalen
    active_upload = db.query(MJOPUpload).filter(
        MJOPUpload.vve_id == vve_id, MJOPUpload.status == "active"
    ).order_by(MJOPUpload.uploaded_at.desc()).first()
    if not active_upload:
        return []
    items = db.query(MJOPItem).filter(MJOPItem.mjop_upload_id == active_upload.id).all()
    for item in items:
        _ = item.quotes  # eager load
    return items


@router.post("/mjop/items", response_model=MJOPItemOut, status_code=status.HTTP_201_CREATED)
def create_mjop_item(
    vve_id: int,
    data: MJOPItemCreate,
    current_user: User = Depends(get_current_beheerder),
    db: Session = Depends(get_db),
):
    _check_vve_access(vve_id, current_user, db)
    active_upload = db.query(MJOPUpload).filter(
        MJOPUpload.vve_id == vve_id, MJOPUpload.status == "active"
    ).order_by(MJOPUpload.uploaded_at.desc()).first()
    if not active_upload:
        # Maak automatisch een "handmatig" uploadrecord aan als er geen actieve bestaat
        active_upload = MJOPUpload(
            vve_id=vve_id,
            original_filename="Handmatig ingevoerd",
            stored_filename="manual",
            status="active",
            uploaded_by_id=current_user.id,
        )
        db.add(active_upload)
        db.commit()
        db.refresh(active_upload)
    item = MJOPItem(vve_id=vve_id, mjop_upload_id=active_upload.id, **data.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.patch("/mjop/items/{item_id}", response_model=MJOPItemOut)
def update_mjop_item(
    vve_id: int,
    item_id: int,
    data: MJOPItemUpdate,
    current_user: User = Depends(get_current_beheerder),
    db: Session = Depends(get_db),
):
    _check_vve_access(vve_id, current_user, db)
    item = db.query(MJOPItem).filter(MJOPItem.id == item_id, MJOPItem.vve_id == vve_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="MJOP post niet gevonden")
    updates = data.model_dump(exclude_unset=True)
    for field, value in updates.items():
        setattr(item, field, value)
    if any(k in updates for k in ("planned_year", "planned_quarter", "planned_amount")):
        item.manually_adjusted = True
    db.commit()
    db.refresh(item)
    return item


# --- MJOP export ---

@router.get("/mjop/export")
def export_mjop(vve_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Exporteer actieve MJOP-posten als Excel-bestand."""
    import io
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    _check_vve_access(vve_id, current_user, db)

    active_upload = db.query(MJOPUpload).filter(
        MJOPUpload.vve_id == vve_id, MJOPUpload.status == "active"
    ).order_by(MJOPUpload.uploaded_at.desc()).first()

    items: list[MJOPItem] = []
    if active_upload:
        items = db.query(MJOPItem).filter(
            MJOPItem.mjop_upload_id == active_upload.id,
            MJOPItem.status != "cancelled",
        ).order_by(MJOPItem.planned_year, MJOPItem.planned_quarter).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MJOP"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    headers = ["Jaar", "Kwartaal", "Omschrijving", "Categorie", "Begroot (€)", "Werkelijk (€)", "Status"]
    col_widths = [8, 10, 45, 16, 14, 14, 14]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    STATUS_NL = {"planned": "Gepland", "quoted": "Offerte", "approved": "Goedgekeurd",
                 "completed": "Afgerond", "cancelled": "Geannuleerd"}

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.planned_year)
        ws.cell(row=row_idx, column=2, value=f"Q{item.planned_quarter}" if item.planned_quarter else "—")
        ws.cell(row=row_idx, column=3, value=item.description)
        ws.cell(row=row_idx, column=4, value=item.category or "—")
        ws.cell(row=row_idx, column=5, value=float(item.planned_amount))
        ws.cell(row=row_idx, column=6, value=float(item.actual_amount) if item.actual_amount else None)
        ws.cell(row=row_idx, column=7, value=STATUS_NL.get(item.status, item.status))

        # Zebra-stripes
        if row_idx % 2 == 0:
            fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = fill

    # Eurobedragen opmaken
    euro_fmt = '€ #,##0'
    for row_idx in range(2, len(items) + 2):
        ws.cell(row=row_idx, column=5).number_format = euro_fmt
        ws.cell(row=row_idx, column=6).number_format = euro_fmt

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=MJOP_export.xlsx"},
    )


# --- Offertes ---

@router.post("/quotes", response_model=QuoteOut, status_code=status.HTTP_201_CREATED)
async def create_quote(
    vve_id: int,
    data: QuoteCreate,
    file: UploadFile = File(None),
    current_user: User = Depends(get_current_beheerder),
    db: Session = Depends(get_db),
):
    _check_vve_access(vve_id, current_user, db)
    item = db.query(MJOPItem).filter(MJOPItem.id == data.mjop_item_id, MJOPItem.vve_id == vve_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="MJOP post niet gevonden")

    doc_path = None
    if file:
        upload_dir = os.path.join(settings.UPLOAD_DIR, "quotes", str(vve_id))
        os.makedirs(upload_dir, exist_ok=True)
        stored_name = f"{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        doc_path = os.path.join(upload_dir, stored_name)
        with open(doc_path, "wb") as f:
            shutil.copyfileobj(file.file, f)

    quote = Quote(vve_id=vve_id, document_path=doc_path, **data.model_dump())
    db.add(quote)
    item.status = "quoted"
    db.commit()
    db.refresh(quote)
    return quote


@router.post("/quotes/{quote_id}/approve", response_model=QuoteOut)
def approve_quote(
    vve_id: int,
    quote_id: int,
    data: QuoteApprove,
    current_user: User = Depends(get_current_beheerder),
    db: Session = Depends(get_db),
):
    _check_vve_access(vve_id, current_user, db)
    quote = db.query(Quote).filter(Quote.id == quote_id, Quote.vve_id == vve_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Offerte niet gevonden")

    quote.is_approved = True
    quote.approved_at = datetime.now(timezone.utc)
    quote.approved_by_id = current_user.id
    if data.final_amount is not None:
        quote.final_amount = data.final_amount

    # Goedgekeurde offerte wordt de actual van de MJOP post
    item = db.query(MJOPItem).filter(MJOPItem.id == quote.mjop_item_id).first()
    if item:
        item.actual_amount = data.final_amount or quote.quoted_amount
        item.status = "approved"

    db.commit()
    db.refresh(quote)
    return quote


# --- Reservefonds ---

@router.get("/reservefonds", response_model=list[ReserveFondsEntryOut])
def list_reservefonds(vve_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _check_vve_access(vve_id, current_user, db)
    return db.query(ReserveFondsEntry).filter(ReserveFondsEntry.vve_id == vve_id).order_by(ReserveFondsEntry.entry_date).all()


@router.post("/reservefonds", response_model=ReserveFondsEntryOut, status_code=status.HTTP_201_CREATED)
def add_reservefonds_entry(
    vve_id: int,
    data: ReserveFondsEntryCreate,
    current_user: User = Depends(get_current_beheerder),
    db: Session = Depends(get_db),
):
    _check_vve_access(vve_id, current_user, db)
    entry = ReserveFondsEntry(vve_id=vve_id, **data.model_dump())
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return entry


@router.get("/reservefonds/combined")
def list_reservefonds_combined(vve_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Geeft alle reservefonds-mutaties terug, inclusief berekende bijdrage-betalingen per periode."""
    _check_vve_access(vve_id, current_user, db)
    from app.models.vve import VvE

    real_entries = (
        db.query(ReserveFondsEntry)
        .filter(ReserveFondsEntry.vve_id == vve_id)
        .order_by(ReserveFondsEntry.entry_date)
        .all()
    )
    plans = (
        db.query(ContributionPlan)
        .filter(ContributionPlan.vve_id == vve_id)
        .order_by(ContributionPlan.effective_from)
        .all()
    )
    vve = db.query(VvE).filter(VvE.id == vve_id).first()
    frequency = vve.contribution_frequency if vve else "monthly"
    today = date.today()

    NL_MONTHS = ["", "januari", "februari", "maart", "april", "mei", "juni",
                 "juli", "augustus", "september", "oktober", "november", "december"]

    def _payment_dates(plan: ContributionPlan, after: date, until: date):
        """Genereer (payment_date, beschrijving) voor alle betalingen van dit plan."""
        results = []
        if frequency == "monthly":
            y, m = after.year, after.month
            while True:
                pd = date(y, m, 20)
                if after < pd <= until:
                    results.append((pd, f"Maandelijkse bijdrage {NL_MONTHS[pd.month]} {pd.year}"))
                m += 1
                if m > 12:
                    y, m = y + 1, 1
                if date(y, m, 1) > until:
                    break
        else:
            for y in range(after.year, until.year + 1):
                for m, q in ((1, 1), (4, 2), (7, 3), (10, 4)):
                    pd = date(y, m, 20)
                    if after < pd <= until:
                        results.append((pd, f"Kwartaalbijdrage Q{q} {y}"))
        return results

    contribution_rows = []
    if real_entries and plans:
        first_entry_date = real_entries[0].entry_date
        for plan in plans:
            start = max(plan.effective_from, first_entry_date)
            end = min(plan.effective_to if plan.effective_to else today, today)
            for pd, desc in _payment_dates(plan, start, end):
                contribution_rows.append({
                    "id": None,
                    "entry_date": pd.isoformat(),
                    "amount": float(plan.amount_per_period),
                    "description": desc,
                    "entry_type": "contribution",
                })

    manual_rows = [
        {
            "id": e.id,
            "entry_date": e.entry_date.isoformat(),
            "amount": float(e.amount),
            "description": e.description or "",
            "entry_type": "manual",
        }
        for e in real_entries
    ]

    all_rows = sorted(manual_rows + contribution_rows, key=lambda r: (r["entry_date"], 0 if r["entry_type"] == "manual" else 1))

    running = 0.0
    for row in all_rows:
        running += row["amount"]
        row["running_balance"] = round(running, 2)

    return all_rows


@router.post("/balanssheet/upload")
async def upload_balanssheet(
    vve_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_beheerder),
    db: Session = Depends(get_db),
):
    """Upload een bankbalans (PDF/Excel) en extraheer het saldo en de datum."""
    _check_vve_access(vve_id, current_user, db)

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in (".pdf", ".xlsx", ".xls", ".xlsm"):
        raise HTTPException(status_code=400, detail="Alleen PDF of Excel bestanden zijn toegestaan.")

    upload_dir = "uploads"
    os.makedirs(upload_dir, exist_ok=True)
    tmp_path = os.path.join(upload_dir, f"balanssheet_{vve_id}{ext}")

    with open(tmp_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    try:
        result = parse_balanssheet(tmp_path)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Bestand kon niet worden verwerkt: {e}")
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    return result


# --- Bijdrageplan ---

@router.get("/contributions", response_model=list[ContributionPlanOut])
def list_contributions(vve_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _check_vve_access(vve_id, current_user, db)
    return db.query(ContributionPlan).filter(ContributionPlan.vve_id == vve_id).order_by(ContributionPlan.effective_from.desc()).all()


@router.post("/contributions", response_model=ContributionPlanOut, status_code=status.HTTP_201_CREATED)
def create_contribution_plan(
    vve_id: int,
    data: ContributionPlanCreate,
    current_user: User = Depends(get_current_beheerder),
    db: Session = Depends(get_db),
):
    _check_vve_access(vve_id, current_user, db)
    # Sluit vorig actief plan automatisch af (effective_to = dag voor nieuwe ingangsdatum)
    prev = db.query(ContributionPlan).filter(
        ContributionPlan.vve_id == vve_id,
        ContributionPlan.effective_to.is_(None),
        ContributionPlan.effective_from < data.effective_from,
    ).order_by(ContributionPlan.effective_from.desc()).first()
    if prev:
        prev.effective_to = data.effective_from - timedelta(days=1)
        db.add(prev)
    plan = ContributionPlan(vve_id=vve_id, **data.model_dump())
    db.add(plan)
    db.commit()
    db.refresh(plan)
    return plan


# --- Dashboard / Scenarios ---

@router.get("/dashboard")
def get_financial_dashboard(
    vve_id: int,
    inflatie: float = Query(default=0.0, ge=0.0, le=20.0),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _check_vve_access(vve_id, current_user, db)
    from app.models.vve import VvE
    from app.models.appartement import Appartement as AppartementModel

    vve = db.query(VvE).filter(VvE.id == vve_id).first()
    if not vve:
        raise HTTPException(status_code=404, detail="VvE niet gevonden")

    today = date.today()

    # Actueel bijdrageplan
    plan = db.query(ContributionPlan).filter(
        ContributionPlan.vve_id == vve_id,
        ContributionPlan.effective_from <= today,
    ).order_by(ContributionPlan.effective_from.desc()).first()
    contribution = plan.amount_per_period if plan else Decimal(0)

    # Huidig reservefonds saldo + automatische bijschrijvingen op de 20ste
    entries = db.query(ReserveFondsEntry).filter(ReserveFondsEntry.vve_id == vve_id).all()
    balance = sum(e.amount for e in entries) if entries else Decimal(0)

    if entries and plan:
        last_date = max(e.entry_date for e in entries)
        balance += _bereken_auto_bijdragen(last_date, today, plan.amount_per_period, vve.contribution_frequency)

    # MJOP items van actief upload
    active_upload = db.query(MJOPUpload).filter(
        MJOPUpload.vve_id == vve_id, MJOPUpload.status == "active"
    ).order_by(MJOPUpload.uploaded_at.desc()).first()

    mjop_items = []
    if active_upload:
        db_items = db.query(MJOPItem).filter(MJOPItem.mjop_upload_id == active_upload.id).all()
        mjop_items = [
            MJOPItemInput(
                id=i.id,
                planned_year=i.planned_year,
                planned_quarter=i.planned_quarter,
                planned_amount=i.planned_amount,
                description=i.description,
            )
            for i in db_items
            if i.status != "cancelled"
        ]

    # Appartementen met aandeel
    appartementen = db.query(AppartementModel).filter(
        AppartementModel.vve_id == vve_id, AppartementModel.is_active == True
    ).order_by(AppartementModel.nummer, AppartementModel.naam).all()
    aandelen = [a.aandeel for a in appartementen]
    totaal_aandeel = sum(aandelen) if aandelen else Decimal(0)

    # Bijdrage per appartement
    share_denominator = vve.share_denominator or 1
    bijdrage_per_eenheid = (
        float(contribution / share_denominator)
        if share_denominator > 1 and contribution > 0
        else None
    )
    bijdrage_per_appartement = [
        {
            "id": a.id,
            "naam": a.naam,
            "nummer": a.nummer,
            "eigenaar_naam": a.eigenaar_naam,
            "aandeel": float(a.aandeel),
            "bijdrage_per_periode": float(
                a.aandeel / totaal_aandeel * contribution if totaal_aandeel > 0 else Decimal(0)
            ),
        }
        for a in appartementen
    ]

    inp = FinancialInput(
        current_balance=balance,
        contribution_per_period=contribution,
        contribution_frequency=vve.contribution_frequency,
        mjop_items=mjop_items,
        current_year=today.year,
        current_quarter=(today.month - 1) // 3 + 1,
        member_aandelen=aandelen,
        inflatie_percentage=Decimal(str(inflatie)),
    )

    result = calculate_all_scenarios(inp)
    result["share_denominator"] = share_denominator
    result["bijdrage_per_eenheid"] = bijdrage_per_eenheid
    result["bijdrage_per_appartement"] = bijdrage_per_appartement
    return result


@router.get("/smart-plan")
def get_smart_plan(
    vve_id: int,
    max_shift_quarters: int = Query(default=8, ge=1, le=16),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _check_vve_access(vve_id, current_user, db)
    from app.models.vve import VvE
    from app.models.appartement import Appartement as AppartementModel

    vve = db.query(VvE).filter(VvE.id == vve_id).first()
    if not vve:
        raise HTTPException(status_code=404, detail="VvE niet gevonden")

    today = date.today()

    plan = db.query(ContributionPlan).filter(
        ContributionPlan.vve_id == vve_id,
        ContributionPlan.effective_from <= today,
    ).order_by(ContributionPlan.effective_from.desc()).first()
    contribution = plan.amount_per_period if plan else Decimal(0)

    entries = db.query(ReserveFondsEntry).filter(ReserveFondsEntry.vve_id == vve_id).all()
    balance = sum(e.amount for e in entries) if entries else Decimal(0)
    if entries and plan:
        last_date = max(e.entry_date for e in entries)
        balance += _bereken_auto_bijdragen(last_date, today, plan.amount_per_period, vve.contribution_frequency)

    active_upload = db.query(MJOPUpload).filter(
        MJOPUpload.vve_id == vve_id, MJOPUpload.status == "active"
    ).order_by(MJOPUpload.uploaded_at.desc()).first()

    mjop_items = []
    if active_upload:
        db_items = db.query(MJOPItem).filter(MJOPItem.mjop_upload_id == active_upload.id).all()
        mjop_items = [
            MJOPItemInput(
                id=i.id,
                planned_year=i.planned_year,
                planned_quarter=i.planned_quarter,
                planned_amount=i.planned_amount,
                description=i.description,
            )
            for i in db_items
            if i.status != "cancelled"
        ]

    appartementen = db.query(AppartementModel).filter(
        AppartementModel.vve_id == vve_id, AppartementModel.is_active == True
    ).all()
    aandelen = [a.aandeel for a in appartementen]

    inp = FinancialInput(
        current_balance=balance,
        contribution_per_period=contribution,
        contribution_frequency=vve.contribution_frequency,
        mjop_items=mjop_items,
        current_year=today.year,
        current_quarter=(today.month - 1) // 3 + 1,
        member_aandelen=aandelen,
    )

    return smart_planning(inp, max_shift_quarters=max_shift_quarters)


@router.post("/mjop/assign-quarters")
def assign_quarters(
    vve_id: int,
    quarter: int = Query(default=1, ge=1, le=4),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Wijs het opgegeven kwartaal toe aan alle MJOP-posten zonder kwartaal (niet-geannuleerd)."""
    _check_vve_access(vve_id, current_user, db)

    active_upload = db.query(MJOPUpload).filter(
        MJOPUpload.vve_id == vve_id, MJOPUpload.status == "active"
    ).order_by(MJOPUpload.uploaded_at.desc()).first()
    if not active_upload:
        raise HTTPException(status_code=404, detail="Geen actief MJOP gevonden")

    items = db.query(MJOPItem).filter(
        MJOPItem.mjop_upload_id == active_upload.id,
        MJOPItem.planned_quarter.is_(None),
        MJOPItem.status != "cancelled",
    ).all()

    for item in items:
        item.planned_quarter = quarter
        item.manually_adjusted = True

    db.commit()
    return {"updated": len(items)}
